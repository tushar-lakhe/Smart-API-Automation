import os
import sys
import ujson as json
import openpyxl as pd
from pathlib import Path
import configparser
import config_url as conf

"""
#ini_path = Path(__file__).parents[7]
#ini_path = os.environ['r_path']

print(r_path)

print(sys.argv)
print("r_path is "+sys.argv[0])


iniFilepath = os.path.join(r_path,'ParameterFile.xlsx').replace('\\', '/')"""

filename = '../../../../../../../../ops/ParameterFile.xlsx'
print(filename)



for k, v in conf.url.items():
    # read parameter excel
    if k == "jsonforkarate":
        featureinputjson = v

# read parameter excel
xl = pd.load_workbook(iniFilepath)
ws = xl['rest_api']
ws1 = xl['message_queue']
first_row = []
dict3 = []
excel_list = []
excel_list1 = []
first_row1 = []

for col in range(1, ws.max_column+1):
    first_row.append(ws.cell(row=1, column=col).value)
excel_list = []

for row in range(2, ws.max_row+1):
    elm = {}
    for col in range(1, ws.max_column+1):
        if first_row[col-1] == "test_case_id":
            elm[first_row[col-1]]="rest_"+str(ws.cell(row=row,column=col).value)
        elif first_row[col-1] == "layered_call":
            elm[first_row[col-1]]="rest_"+str(ws.cell(row=row,column=col).value)
        elif first_row[col-1] == "json_payload":
            elm[first_row[col-1]]=str(ws.cell(row=row,column=col).value).replace('\n', ' ').replace('\r', '')
        else:
            elm[first_row[col-1]]=ws.cell(row=row,column=col).value

    excel_list.append(elm)

##############################################################################

for col in range(1, ws1.max_column+1):
    first_row1.append(ws1.cell(row=1, column=col).value)

for row in range(2, ws1.max_row+1):
    elm1 = {}
    for col in range(1, ws1.max_column+1):
        if first_row1[col-1] == "test_case_id":
            elm1[first_row1[col-1]]="ems_"+str(ws1.cell(row=row,column=col).value)
        elif first_row1[col-1] == "json_payload":
            elm1[first_row1[col-1]]=str(ws1.cell(row=row,column=col).value).replace('\n', ' ').replace('\r', '')
        else:
            elm1[first_row1[col-1]]=ws1.cell(row=row,column=col).value

    excel_list.append(elm1)

############################################################################

for keys in excel_list:
    dict3.append(keys)

with open(featureinputjson, "w", encoding="utf-8") as writeJsonfile:
    json.dump(dict3, writeJsonfile, indent=4)
